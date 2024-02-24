import numpy as np
import read
from GA import GA
import openpyxl


def test_dataset(dataset=r'knapPI_1_50_1000.csv'):
    fitness_end_d = 1  # 迭代终止判断条件，若最后十次适应度标准差小于此值则终止迭代
    n_ex, c_ex, z_ex, value_ex, wei_ex, code_ex = read.read_csv(dataset)  # 读取文件
    print(len(n_ex))  # 打印问题数量
    suc_g = 0  # 记录总共运行结果等于最优结果次数用来求出最后的数据集正确率
    generation_g = 0  # 记录总共迭代次数
    m = 5  # 每个问题重复运行次数

    # 检查文件是否存在，如果不存在则创建
    file_name = "result.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # 获取当前活动的工作表
    ws = workbook.active

    # 获取当前文件最后一行
    now_row = ws.max_row
    new_row = ws.max_row + 1

    for ex_i in range(len(n_ex)):  # 开始循环解题
        max_weight = c_ex[ex_i]  # 获取最大限制重量
        values = value_ex[ex_i]  # 获取每个物品价值
        weights = wei_ex[ex_i]   # 获取每个物品重量

        answer_ex = z_ex[ex_i]   # 获取最优解
        # answer_code_ex = code_ex[ex_i]  # 获取最优结果基因序列

        def qualified(qualified_testcode):  # 定义贪婪算法使染色体合格化，输入染色体;输出染色体
            qualified_testcode = np.array(qualified_testcode)  # 导入染色体记录基因序列
            qualified_total_weight = sum(weights[qualified_testcode == 1])  # 计算所有基因为1的物品重量总和
            while qualified_total_weight > max_weight:  # 如果超过最大限制重量，随机丢弃至满足条件为止
                indices_of_ones = np.flatnonzero(qualified_testcode == 1)  # 获取染色体基因为1的位置
                random_index = np.random.choice(indices_of_ones, size=1)[0]  # 随机选取一个为1的基因
                qualified_testcode[random_index] = 0  # 随机丢弃
                qualified_total_weight = sum(weights[qualified_testcode == 1])  # 计算所有基因为1的物品重量总和
            return qualified_testcode

        def fitness_func_01packet(testcode):
            testcode = np.array(testcode)  # 导入染色体记录基因序列
            # 随机丢弃后染色体符合不超过最大限制重量条件，计算适应度
            total_value = sum(values[testcode == 1])  # 计算所有基因为1的物品价值总和
            fitness = total_value  # 适应度即为总价值
            return fitness  # 返回适应度值

        if now_row == 1:  # 当结果表格文件无内容时
            suc = 0  # 初始化当前问题中运行结果等于最优结果次数
            result_ex = np.array([])  # 初始化染色体
            generation = 0  # 初始化当前问题总迭代次数
            for i in range(m):  # 对当前进行m次遗传算法
                # ---------------------------------------------------
                # 导入参数：适应度评估函数、编码长度、种群个体数量、终止条件参数
                ga_test = GA(fitness_func=fitness_func_01packet, qualified_func=qualified
                             , n=len(weights), population_size=50,
                             fitness_d=fitness_end_d)
                # ---------------------------------------------------
                # 获取运行结果：结果染色体基因序列、结果总价值（即种群适应度）
                optimized_code, optimized_result, optimized_generation = ga_test.run()
                # 将染色体基因序列加入result_ex列表中
                result_ex = np.append(result_ex, optimized_result)
                # 如果运行结果总价值等于最优结果总价值
                if optimized_result == answer_ex:
                    suc = suc + 1  # 当前问题中运行结果等于最优结果次数加一
                    suc_g = suc_g + 1  # 总共运行结果等于最优结果次数加一
                # 计入迭代次数
                generation = generation + optimized_generation  # 当前问题
                generation_g = generation_g + optimized_generation  # 当前数据集
            # 打印当前问题正确率、平均迭代次数
            print("第 ", ex_i + 1, " 个例子正确率：", suc / m, " ,平均迭代次数：", generation / m)
            # 打印当前问题误差百分比，运行m次结果标准差
            print("第 ", ex_i + 1, " 个例子的误差百分比：", abs(answer_ex-np.mean(result_ex))/answer_ex*100,
                  "% ,标准差为：", np.std(result_ex))
            # 将结果写入Excel文件
            ws.cell(row=1, column=4 * ex_i + 1, value="第" + str(ex_i + 1) + "例正确率")  # 初始化第一行正确率
            ws.cell(row=2, column=4 * ex_i + 1, value=suc / m)  # 填入正确率
            ws.cell(row=1, column=4 * ex_i + 2, value="平均迭代次数")  # 初始化第一行平均迭代次数
            ws.cell(row=2, column=4 * ex_i + 2, value=generation / m)  # 填入平均迭代次数
            ws.cell(row=1, column=4 * ex_i + 3, value="误差百分比")  # 初始化第一行误差百分比
            ws.cell(row=2, column=4 * ex_i + 3, value=str(abs(answer_ex-np.mean(result_ex))/answer_ex*100) + "%")
            # 填入误差百分比
            ws.cell(row=1, column=4 * ex_i + 4, value="标准差")  # 初始化第一行标准差
            ws.cell(row=2, column=4 * ex_i + 4, value=np.std(result_ex))  # 填入标准差
        else:  # 当已经有结果填入结果表格文件时
            suc = 0  # 初始化当前问题中运行结果等于最优结果次数
            result_ex = np.array([])  # 初始化染色体
            generation = 0  # 初始化当前问题总迭代次数
            for i in range(m):  # 对当前进行m次遗传算法
                # ---------------------------------------------------
                # 导入参数：适应度评估函数、编码长度、种群个体数量、终止条件参数
                ga_test = GA(fitness_func=fitness_func_01packet, qualified_func=qualified
                             , n=len(weights), population_size=50,
                             fitness_d=fitness_end_d)
                # ---------------------------------------------------
                # 获取运行结果：结果染色体基因序列、结果总价值（即种群适应度）、迭代次数
                optimized_code, optimized_result, optimized_generation = ga_test.run()
                # 将染色体基因序列加入result_ex列表中
                result_ex = np.append(result_ex, optimized_result)
                # 如果运行结果总价值等于最优结果总价值
                if optimized_result == answer_ex:
                    suc = suc + 1  # 当前问题中运行结果等于最优结果次数加一
                    suc_g = suc_g + 1  # 总共运行结果等于最优结果次数加一
                # 计入迭代次数
                generation = generation + optimized_generation  # 当前问题
                generation_g = generation_g + optimized_generation  # 当前数据集
            print("第 ", ex_i + 1, " 个例子正确率：", suc / m, " ,平均迭代次数：", generation / m)  # 打印当前问题正确率、平均迭代次数
            print("第 ", ex_i + 1, " 个例子的误差百分比：", abs(answer_ex - np.mean(result_ex)) / answer_ex * 100,
                  "% ,标准差为：", np.std(result_ex))  # 打印当前问题误差百分比，运行m次结果标准差
            # 将结果写入Excel文件
            ws.cell(row=new_row, column=4 * ex_i + 1, value=suc / m)  # 填入正确率
            ws.cell(row=new_row, column=4 * ex_i + 2, value=generation / m)  # 填入平均迭代次数
            ws.cell(row=new_row, column=4 * ex_i + 3, value=str(abs(answer_ex-np.mean(result_ex))/answer_ex*100) + "%")
            # 填入误差百分比
            ws.cell(row=new_row, column=4 * ex_i + 4, value=np.std(result_ex))  # 填入标准差

    print("数据集正确率：", suc_g / (m*len(n_ex)), " ,平均迭代次数", generation_g / (m*len(n_ex)))  # 打印数据集正确率和平均迭代次数
    if now_row == 1:  # 当结果表格文件无内容时
        ws.cell(row=1, column=4 * (len(n_ex) - 1) + 5, value="数据集正确率")  # 初始化第一行数据集正确率
        ws.cell(row=2, column=4 * (len(n_ex) - 1) + 5, value=suc_g / (m * len(n_ex)))  # 填入数据集正确率
        ws.cell(row=1, column=4 * (len(n_ex) - 1) + 6, value="平均迭代次数")  # 初始化第一行数据集平均迭代次数
        ws.cell(row=2, column=4 * (len(n_ex) - 1) + 6, value=generation_g / (m*len(n_ex)))  # 填入数据集平均迭代次数
    else:  # 当已经有结果填入结果表格文件时
        ws.cell(row=new_row, column=4 * (len(n_ex) - 1) + 5, value=suc_g / (m * len(n_ex)))  # 填入数据集正确率
        ws.cell(row=new_row, column=4 * (len(n_ex) - 1) + 6, value=generation_g / (m*len(n_ex)))  # 填入数据集平均迭代次数

    # 保存工作簿到文件
    workbook.save(file_name)

    # 关闭工作簿
    workbook.close()


if __name__ == "__main__":
    test_dataset(dataset=r'knapPI_1_50_1000.csv')  # 读取数据文件
